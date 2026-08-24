# -*- coding: utf-8 -*-
"""华信数智房产交易一体化平台 菜单梳理表生成器

读《柳州市数智房产交易一体化平台功能板块梳理v1.9.xlsx》与 tools/menu_spec.py，
生成《柳州市数智房产交易一体化平台菜单梳理v1.0.xlsx》，四个 sheet：

  子系统一览   业务子系统清单与规模
  菜单明细     一级菜单 / 二级菜单逐条，附对应的 v1.9 功能板块与一级功能模块
  功能点归属   v1.9 全部二级功能点逐条落到具体菜单，用于验证不重不漏
  变更对照     v1.9 功能板块 / 一级功能模块 → 新菜单，标注保留 / 改名 / 合并 / 拆分

生成前跑四项硬校验，任一不通过即中止：
  1. 各子系统承接板块下的一级功能模块与二级功能点全部有归属，无遗漏、无重复
  2. 一级菜单名不超过 MAX_L1 字、二级菜单名不超过 MAX_L2 字
  3. 同一子系统内菜单名不重复
  4. 每个子系统的第一项都是「我的工作台」

只在梳理期运行，不参与页面运行。用法：在 prototype/ 下执行 python tools/gen_menu.py
"""

import os
import sys
from collections import OrderedDict

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from menu_spec import MAX_L1, MAX_L2, SKIPPED_DOMAINS, SUBSYSTEMS  # noqa: E402

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DOCS = os.path.dirname(ROOT)
XLSX_IN = os.path.join(DOCS, '柳州市数智房产交易一体化平台功能板块梳理v1.9.xlsx')
XLSX_OUT = os.path.join(DOCS, '柳州市数智房产交易一体化平台菜单梳理v1.0.xlsx')

SHEET_END = [('2PC端应用', 'government'), ('1数据中心', 'datacenter'),
             ('3移动端应用', 'mobile'), ('4AI应用', 'ai')]
END_NAME = {'government': '业务办理端', 'datacenter': '数据中心', 'ai': '数智大脑',
            'portal': '统一服务门户', 'mobile': '移动端'}

# 2PC端应用 sheet 里混着对外门户的功能域，按功能域改判所属端
DOMAIN_END = {'统一服务门户': 'portal'}


def norm(v):
    return '' if v is None else str(v).strip()


def width(s):
    """显示宽度：一个汉字算 1，一个 ASCII 字符算 0.5。"""
    return sum(1.0 if ord(c) > 0x2E80 else 0.5 for c in s)


# ---------------------------------------------------------------------------
# 一、读 v1.9 梳理表
# ---------------------------------------------------------------------------
def load_v19():
    wb = openpyxl.load_workbook(XLSX_IN, data_only=True)

    intro, legacy = {}, {}
    ws = wb['功能板块一览']
    ci = {norm(c.value): i for i, c in enumerate(ws[1])}
    for r in ws.iter_rows(min_row=2, values_only=True):
        name = norm(r[ci['功能板块']])
        if name:
            intro[name] = norm(r[ci['板块定位与建设内容']])
            legacy[name] = norm(r[ci['承接原有子系统']])

    boards = OrderedDict()
    for sheet, end in SHEET_END:
        ws = wb[sheet]
        ci = {norm(c.value): i for i, c in enumerate(ws[1])}
        domain = board = module = ''
        for r in ws.iter_rows(min_row=2, values_only=True):
            # 4AI应用 表尾附有「AI 能力赋能对照」小表，序号以「附」开头，不是功能板块
            if norm(r[ci['序号']]).startswith('附'):
                break
            domain = norm(r[ci['功能域']]) or domain
            board = norm(r[ci['功能板块']]) or board
            module = norm(r[ci['一级功能模块']]) or module
            point = norm(r[ci['二级功能点']])
            if not point:
                continue
            b = boards.setdefault(board, {
                'domain': domain, 'end': DOMAIN_END.get(domain, end), 'name': board,
                'intro': intro.get(board, ''), 'legacy': legacy.get(board, ''),
                'modules': OrderedDict(),
            })
            # 板块名是全表主键，重名会让两个 sheet 的模块串在一起，必须先发现
            if b['domain'] != domain:
                raise SystemExit('板块名「%s」在功能域「%s」与「%s」下重复出现'
                                 % (board, b['domain'], domain))
            b['modules'].setdefault(module, []).append({
                'name': point,
                'desc': norm(r[ci['功能说明']]),
                'level': norm(r[ci['适用层级']]),
                'obj': norm(r[ci['主要使用对象']]),
                'ai': norm(r[ci['是否 AI 赋能']]),
                'cfg': norm(r[ci['一城一策可配置']]),
            })
    return boards


# ---------------------------------------------------------------------------
# 二、展开菜单树
# ---------------------------------------------------------------------------
def build_menu(boards):
    """把 SUBSYSTEMS 展开成扁平的菜单项列表，并把功能点挂到菜单项上。"""
    items = []          # 每项：子系统 / 一级菜单 / 二级菜单 / 菜单代号 / 来源 / 功能点
    errors = []

    for sub in SUBSYSTEMS:
        seq = 0
        for gname, entries in sub['menu']:
            for label, srcs in entries:
                seq += 1
                pts, src_boards, src_modules = [], [], []
                for src in srcs:
                    bname, mname = src[0], src[1]
                    b = boards.get(bname)
                    if b is None:
                        errors.append('%s > %s：板块「%s」在 v1.9 中不存在' % (sub['name'], label, bname))
                        continue
                    if mname not in b['modules']:
                        errors.append('%s > %s：板块「%s」下没有模块「%s」'
                                      % (sub['name'], label, bname, mname))
                        continue
                    all_pts = b['modules'][mname]
                    if len(src) == 3:
                        want = list(src[2])
                        have = [p['name'] for p in all_pts]
                        for w in want:
                            if w not in have:
                                errors.append('%s > %s：模块「%s」下没有功能点「%s」'
                                              % (sub['name'], label, mname, w))
                        take = [p for p in all_pts if p['name'] in want]
                    else:
                        take = all_pts
                    pts.extend(take)
                    if bname not in src_boards:
                        src_boards.append(bname)
                    if mname not in src_modules:
                        src_modules.append(mname)
                items.append({
                    'sub': sub, 'group': gname, 'label': label,
                    'key': '%s-%02d' % (sub['key'], seq),
                    'srcs': srcs, 'src_boards': src_boards, 'src_modules': src_modules,
                    'points': pts,
                })
        # 重名检查按「组内 + 顶层」范围做：按服务/环节分组的子系统（如统一应用服务平台）
        # 会在多个组下各设「应用配置/参数配置/服务调用日志/接口调试」等同名菜单，
        # 它们分属不同一级菜单、菜单代号不同，属正常设计，只在同一组内或顶层单项之间判重。
        top = []
        for gname, entries in sub['menu']:
            names = [lbl for lbl, _ in entries]
            if gname:
                top.append(gname)
                gdup = sorted({x for x in names if names.count(x) > 1})
                if gdup:
                    errors.append('%s > %s：同组菜单名重复 %s' % (sub['name'], gname, '、'.join(gdup)))
            else:
                top.extend(names)
        tdup = sorted({x for x in top if top.count(x) > 1})
        if tdup:
            errors.append('%s：一级/顶层菜单名重复 %s' % (sub['name'], '、'.join(tdup)))
    return items, errors


# ---------------------------------------------------------------------------
# 三、四项硬校验
# ---------------------------------------------------------------------------
def validate(boards, items):
    errors = []

    # 1. 覆盖度：承接板块下的每个功能点恰好归属一次
    covered = {}
    for it in items:
        for src in it['srcs']:
            bname, mname = src[0], src[1]
            if bname not in boards or mname not in boards[bname]['modules']:
                continue
            names = list(src[2]) if len(src) == 3 else [p['name'] for p in boards[bname]['modules'][mname]]
            for n in names:
                covered.setdefault((bname, mname, n), []).append(it['sub']['name'] + ' > ' + it['label'])

    claimed = set()
    for sub in SUBSYSTEMS:
        for b in sub['boards']:
            if b not in boards:
                errors.append('%s：承接板块「%s」在 v1.9 中不存在' % (sub['name'], b))
            elif b in claimed:
                errors.append('板块「%s」被多个子系统重复承接' % b)
            else:
                claimed.add(b)

    for bname, b in boards.items():
        if bname not in claimed:
            if b['domain'] not in SKIPPED_DOMAINS:
                errors.append('板块「%s」（功能域 %s）没有任何子系统承接' % (bname, b['domain']))
            continue
        for mname, pts in b['modules'].items():
            for p in pts:
                owners = covered.get((bname, mname, p['name']), [])
                if not owners:
                    errors.append('功能点未归属：%s > %s > %s' % (bname, mname, p['name']))
                elif len(owners) > 1:
                    errors.append('功能点重复归属：%s > %s > %s → %s'
                                  % (bname, mname, p['name'], '，'.join(owners)))

    # 2. 菜单名长度
    for sub in SUBSYSTEMS:
        for gname, entries in sub['menu']:
            if gname and width(gname) > MAX_L1:
                errors.append('一级菜单名超长（上限 %g）：%s > %s' % (MAX_L1, sub['name'], gname))
            for label, _ in entries:
                limit = MAX_L1 if not gname else MAX_L2
                if width(label) > limit:
                    errors.append('菜单名超长（上限 %g）：%s > %s' % (limit, sub['name'], label))

    # 3. 同一子系统内菜单名不重复 —— 已在 build_menu 中检查

    # 4. 每个子系统的第一项都是「我的工作台」
    for sub in SUBSYSTEMS:
        first = sub['menu'][0][1][0][0] if sub['menu'] and sub['menu'][0][1] else ''
        if first != '我的工作台':
            errors.append('%s：第一项不是「我的工作台」，实际是「%s」' % (sub['name'], first))

    return errors


# ---------------------------------------------------------------------------
# 四、写 Excel
# ---------------------------------------------------------------------------
HEAD_FILL = PatternFill('solid', fgColor='2656BB')
HEAD_FONT = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
BODY_FONT = Font(name='微软雅黑', size=10)
GROUP_FONT = Font(name='微软雅黑', size=10, bold=True, color='1F3D7A')
THIN = Side(style='thin', color='C9D2E3')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOP_WRAP = Alignment(vertical='top', wrap_text=True)
TOP_LEFT = Alignment(vertical='top', horizontal='left')
CENTER = Alignment(vertical='center', horizontal='center')
MERGED_ALIGN = Alignment(vertical='center', horizontal='left', wrap_text=True)


def write_sheet(wb, title, header, widths, rows, first=False):
    ws = wb.active if first else wb.create_sheet()
    ws.title = title
    ws.append(header)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for c in ws[1]:
        c.fill, c.font, c.alignment, c.border = HEAD_FILL, HEAD_FONT, CENTER, BORDER
    ws.row_dimensions[1].height = 26
    for row in rows:
        ws.append(row)
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.font = BODY_FONT
            c.border = BORDER
            c.alignment = TOP_WRAP if isinstance(c.value, str) and len(c.value) > 14 else TOP_LEFT
    ws.freeze_panes = 'A2'
    return ws


def merge_sheet(ws, plan):
    """按 plan 纵向合并重复值。plan = [(列号, (限定列号...)), ...]。

    限定列是层级上更靠外、或能唯一标识本行归属的列。带上它，同名的下级
    （比如两个子系统都有「房源超市」一级菜单）才不会被误当成一段连续区间。
    合并前先快照全表取值：merge_cells 之后区间内非首行会读成 None，
    直接读单元格会让后面几列的边界判断失真。
    """
    last = ws.max_row
    grid = {r: [c.value for c in ws[r]] for r in range(2, last + 1)}

    def val(r, col):
        row = grid[r]
        return row[col - 1] if col <= len(row) else None

    for col, guards in plan:
        start, prev = 2, None
        for r in range(2, last + 2):
            cur = None
            if r <= last:
                cur = (tuple(val(r, g) for g in guards), val(r, col))
            if cur != prev:
                if prev is not None and r - start > 1:
                    ws.merge_cells(start_row=start, start_column=col,
                                   end_row=r - 1, end_column=col)
                start, prev = r, cur
        for r in range(2, last + 1):
            c = ws.cell(r, col)
            if not isinstance(c, MergedCell):
                c.alignment = MERGED_ALIGN


def build_workbook(boards, items):
    wb = openpyxl.Workbook()

    # ---- sheet 1 子系统一览 ----
    rows = []
    for i, sub in enumerate(SUBSYSTEMS, 1):
        mine = [it for it in items if it['sub'] is sub]
        l1 = [g for g, _ in sub['menu'] if g]
        top = sum(1 for g, es in sub['menu'] if not g for _ in es)
        l2 = sum(1 for it in mine if it['group'])
        mods = sum(len(boards[b]['modules']) for b in sub['boards'] if b in boards)
        pts = sum(len(p) for b in sub['boards'] if b in boards
                  for p in boards[b]['modules'].values())
        legacy = []
        for b in sub['boards']:
            for x in (boards.get(b, {}).get('legacy', '') or '').split('；'):
                x = x.strip()
                if x and x != '—' and x not in legacy:
                    legacy.append(x)
        rows.append([
            i, sub['name'], sub['key'], END_NAME[sub['end']], sub['line'],
            len(l1) + top, l2, len(mine),
            '；'.join(sub['boards']), mods, pts, '；'.join(legacy) or '—',
        ])
    ws = write_sheet(wb, '子系统一览',
                     ['序号', '子系统名称', '子系统代号', '所属端', '业务线',
                      '一级菜单数', '二级菜单数', '菜单页面数',
                      '承接 v1.9 功能板块', '一级功能模块数', '功能点数', '承接原有子系统'],
                     [6, 30, 11, 12, 10, 11, 11, 11, 46, 14, 10, 46],
                     rows, first=True)
    merge_sheet(ws, [(4, ()), (5, (4,))])       # 所属端 → 业务线

    # ---- sheet 2 菜单明细 ----
    rows = []
    for it in items:
        sub = it['sub']
        l1 = it['group'] or it['label']
        l2 = it['label'] if it['group'] else ''
        name = l2 or l1
        rows.append([
            sub['name'], sub['key'], l1, l2, it['key'], '%g' % width(name),
            '；'.join(it['src_boards']) or '—（新增）',
            '；'.join(it['src_modules']) or '—（新增）',
            len(it['points']),
            '按功能点拆分' if any(len(s) == 3 for s in it['srcs'])
            else ('多模块合并' if len(it['src_modules']) > 1
                  else ('新增菜单' if not it['srcs'] else '')),
        ])
    ws = write_sheet(wb, '菜单明细',
                     ['子系统', '子系统代号', '一级菜单', '二级菜单', '菜单代号', '菜单名字数',
                      '对应 v1.9 功能板块', '对应 v1.9 一级功能模块', '功能点数', '备注'],
                     [28, 11, 14, 14, 12, 11, 30, 30, 10, 14],
                     rows)
    for r in ws.iter_rows(min_row=2, min_col=3, max_col=4):
        if not r[1].value:                      # 顶层单项：一级菜单加粗
            r[0].font = GROUP_FONT
    merge_sheet(ws, [(1, ()), (2, (1,)), (3, (1,))])    # 子系统 → 代号 / 一级菜单

    # ---- sheet 3 功能点归属 ----
    owner = {}
    for it in items:
        for src in it['srcs']:
            bname, mname = src[0], src[1]
            names = (list(src[2]) if len(src) == 3
                     else [p['name'] for p in boards.get(bname, {}).get('modules', {}).get(mname, [])])
            for n in names:
                owner[(bname, mname, n)] = it
    rows = []
    for bname, b in boards.items():
        for mname, pts in b['modules'].items():
            for p in pts:
                it = owner.get((bname, mname, p['name']))
                if it is None:
                    rows.append([b['domain'], bname, mname, p['name'], p['desc'],
                                 '—', '—', '—', '—',
                                 SKIPPED_DOMAINS.get(b['domain'], '未归属')])
                    continue
                sub = it['sub']
                rows.append([b['domain'], bname, mname, p['name'], p['desc'],
                             sub['name'], it['group'] or it['label'],
                             it['label'] if it['group'] else '', it['key'], ''])
    ws = write_sheet(wb, '功能点归属',
                     ['v1.9 功能域', 'v1.9 功能板块', 'v1.9 一级功能模块', 'v1.9 二级功能点',
                      '功能说明', '归属子系统', '一级菜单', '二级菜单', '菜单代号', '备注'],
                     [22, 26, 22, 22, 52, 28, 14, 14, 12, 28],
                     rows)
    # 二级菜单为空串的顶层单项要靠菜单代号（第 9 列）区分，否则会连着合并
    merge_sheet(ws, [(1, ()), (2, (1,)), (3, (1, 2)), (6, (1, 2)),
                     (7, (1, 2, 6)), (9, (1, 2, 6, 7)), (8, (1, 2, 6, 7, 9))])

    # ---- sheet 4 变更对照 ----
    rows = []
    for bname, b in boards.items():
        for mname, pts in b['modules'].items():
            hits = [it for it in items
                    if any(s[0] == bname and s[1] == mname for s in it['srcs'])]
            if not hits:
                rows.append([b['domain'], bname, mname, len(pts), '—', '—',
                             '本轮不梳理', SKIPPED_DOMAINS.get(b['domain'], '')])
                continue
            sub = hits[0]['sub']
            paths = ['%s > %s' % (h['group'], h['label']) if h['group'] else h['label']
                     for h in hits]
            if len(hits) > 1:
                kind = '拆分'
                note = '按功能点拆为 %d 条菜单' % len(hits)
            elif len(hits[0]['src_modules']) > 1:
                kind = '合并'
                note = '与「%s」合并为一条菜单' % '」「'.join(
                    m for m in hits[0]['src_modules'] if m != mname)
            elif hits[0]['label'] == mname:
                kind = '保留'
                note = ''
            else:
                kind = '改名'
                note = '%s → %s' % (mname, hits[0]['label'])
            rows.append([b['domain'], bname, mname, len(pts),
                         sub['name'], '；'.join(paths), kind, note])
    for sub in SUBSYSTEMS:
        for it in items:
            if it['sub'] is sub and not it['srcs']:
                rows.append(['—', '—', '—', 0, sub['name'],
                             '%s > %s' % (it['group'], it['label']) if it['group'] else it['label'],
                             '新增', 'v1.9 中无对应模块，各子系统统一入口'])
    ws = write_sheet(wb, '变更对照',
                     ['v1.9 功能域', 'v1.9 功能板块', 'v1.9 一级功能模块', '功能点数',
                      '归属子系统', '新菜单路径', '变更类型', '说明'],
                     [22, 26, 22, 10, 28, 32, 11, 40],
                     rows)
    merge_sheet(ws, [(1, ()), (2, (1,)), (5, (1, 2))])
    return wb


# ---------------------------------------------------------------------------
# 五、主流程
# ---------------------------------------------------------------------------
def main():
    boards = load_v19()
    items, errors = build_menu(boards)
    errors += validate(boards, items)

    if errors:
        print('校验未通过，共 %d 条：' % len(errors))
        for e in errors:
            print('  - ' + e)
        sys.exit(1)

    wb = build_workbook(boards, items)
    wb.save(XLSX_OUT)

    print('已生成 %s' % os.path.basename(XLSX_OUT))
    print('-' * 76)
    print('%-28s %5s %5s %5s %5s %5s' % ('子系统', '一级', '二级', '页面', '模块', '功能点'))
    for sub in SUBSYSTEMS:
        mine = [it for it in items if it['sub'] is sub]
        l1 = sum(1 for g, es in sub['menu'] for _ in ([g] if g else es))
        l2 = sum(1 for it in mine if it['group'])
        mods = sum(len(boards[b]['modules']) for b in sub['boards'])
        pts = sum(len(p) for b in sub['boards'] for p in boards[b]['modules'].values())
        print('%-28s %5d %5d %5d %5d %5d' % (sub['name'], l1, l2, len(mine), mods, pts))
    print('-' * 76)
    tot_mods = sum(len(boards[b]['modules']) for s in SUBSYSTEMS for b in s['boards'])
    tot_pts = sum(len(p) for s in SUBSYSTEMS for b in s['boards']
                  for p in boards[b]['modules'].values())
    print('%-28s %5s %5s %5d %5d %5d'
          % ('合计 %d 个子系统' % len(SUBSYSTEMS), '', '', len(items), tot_mods, tot_pts))
    skipped = [(b, x['domain']) for b, x in boards.items()
               if x['domain'] in SKIPPED_DOMAINS]
    if skipped:
        print('本轮不梳理：%d 个板块（%s）'
              % (len(skipped), '、'.join(sorted({d for _, d in skipped}))))


if __name__ == '__main__':
    main()
